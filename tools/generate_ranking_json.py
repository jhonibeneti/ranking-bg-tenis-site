import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

SOURCE_URL = 'https://docs.google.com/spreadsheets/d/12WU3GH6SdVNiTok_I_TUcaMQWcHf3YYuD5qFHl9W_UE/htmlview'
XLSX_PATH = Path('/home/ubuntu/tenis-ranking/ranking.xlsx')
OUTPUT_PATH = Path(__file__).resolve().parents[1] / 'public/data/ranking.json'

NAME_FIXES = {
    'BEY': 'BYE',
    'Mathues Klaus': 'Matheus Klaus',
    'Feilipe de Cont': 'Filipe de Conto',
    'Marcelo da Siva Ros': 'Marcelo da Silva Ros',
    'Leonardo Ravanelo': 'Leonardo Ravanello',
    'Rafael Ravanello]': 'Rafael Ravanello',
}

ROUND_COLUMNS = [
    ('C', 'D'), ('F', 'G'), ('I', 'J'), ('L', 'M'), ('O', 'P'),
    ('R', 'S'), ('U', 'V'), ('X', 'Y'), ('AA', 'AB'),
]
ROUND_META = [
    (1, 'Rodada 1', '17/08 — 30/08'),
    (2, 'Rodada 2', '31/08 — 13/09'),
    (3, 'Rodada 3', '14/09 — 27/09'),
    (4, 'Rodada 4', '28/09 — 11/10'),
    (5, 'Rodada 5', '12/10 — 25/10'),
    (6, 'Rodada 6', '26/10 — 08/11'),
    (7, 'Rodada 7', '09/11 — 22/11'),
    (8, 'Rodada 8', '23/11 — 06/12'),
    (9, 'Rodada 9', '07/12 — 21/12'),
]


def clean(value):
    return re.sub(r'\s+', ' ', str(value or '')).strip()


def normalize_name(value):
    name = clean(value).replace('×', '').strip()
    return NAME_FIXES.get(name, name)


def color_is_green(color):
    if not color:
        return False
    if color.type == 'rgb' and color.rgb:
        rgb = color.rgb[-6:].upper()
        try:
            r, g, b = int(rgb[:2], 16), int(rgb[2:4], 16), int(rgb[4:], 16)
            return g > r and g > b and g < 190
        except ValueError:
            return False
    return False


def cell_runs(cell):
    value = cell.value
    if value is None:
        return []
    if isinstance(value, str):
        return [{'text': value, 'bold': bool(cell.font.bold), 'green': color_is_green(cell.font.color)}]
    runs = []
    for block in value:
        text = getattr(block, 'text', '')
        font = getattr(block, 'font', None)
        runs.append({
            'text': text,
            'bold': bool(getattr(font, 'b', False)),
            'green': color_is_green(getattr(font, 'color', None)),
        })
    return runs


def parse_match(cell, score_cell):
    runs = cell_runs(cell)
    if not runs:
        return None
    left_text, right_text = [], []
    left_name_marked = right_name_marked = False
    blank_right_marker = False
    side = 'left'
    for run in runs:
        text = run['text']
        if '×' in text:
            before, after = text.split('×', 1)
            left_text.append(before)
            side = 'right'
            right_text.append(after)
        else:
            (left_text if side == 'left' else right_text).append(text)
        # Google Sheets exports the green/bold winner either on the name itself
        # or as a zero-width/blank rich-text run immediately before the name.
        if run['bold'] and run['green']:
            if text.strip():
                if side == 'left':
                    left_name_marked = True
                else:
                    right_name_marked = True
            elif side == 'right':
                blank_right_marker = True
    a = normalize_name(''.join(left_text))
    b = normalize_name(''.join(right_text))
    if not a or not b:
        return None

    score_raw = clean(score_cell.value)
    tokens = re.findall(r'(\d+)\s*x\s*(\d+)', score_raw, flags=re.IGNORECASE)
    score_pairs = [(int(x), int(y)) for x, y in tokens]
    has_nonzero = any(x != 0 or y != 0 for x, y in score_pairs)
    while score_pairs and score_pairs[-1] == (0, 0) and has_nonzero:
        score_pairs.pop()
    score = '  ·  '.join(f'{x}–{y}' for x, y in score_pairs)

    winner = a if left_name_marked else b if right_name_marked else b if blank_right_marker else None
    if winner == 'BYE':
        winner = b if a == 'BYE' else a
    if a == 'BYE' or b == 'BYE':
        status = 'bye'
        score = '—' if not has_nonzero else score
        winner = None
    elif not has_nonzero or not winner:
        status = 'scheduled'
        winner = None
    else:
        status = 'final'

    result = {'a': a, 'b': b, 'status': status}
    if score:
        result['score'] = score
    if winner and winner != 'BYE' and status == 'final':
        result['winner'] = winner
    return result


def group_data(ws, name, player_rows, match_rows):
    players = [normalize_name(ws[f'A{row}'].value) for row in player_rows if clean(ws[f'A{row}'].value)]
    matches = {}
    for round_id, (match_col, score_col) in enumerate(ROUND_COLUMNS, start=1):
        round_matches = []
        for row in match_rows:
            parsed = parse_match(ws[f'{match_col}{row}'], ws[f'{score_col}{row}'])
            if parsed:
                round_matches.append(parsed)
        if round_matches:
            matches[str(round_id)] = round_matches
    return {'name': name, 'players': players, 'matches': matches}


wb = load_workbook(XLSX_PATH, data_only=False, rich_text=True)
ws = wb[wb.sheetnames[0]]
data = {
    'schemaVersion': 1,
    'updatedAt': datetime.now(timezone.utc).isoformat(),
    'sourceUrl': SOURCE_URL,
    'rounds': [{'id': i, 'label': label, 'dates': dates} for i, label, dates in ROUND_META],
    'groups': [
        group_data(ws, 'Grupo 1', range(10, 19), [10, 12, 14, 16, 18]),
        group_data(ws, 'Grupo 2', range(24, 33), [24, 26, 28, 30, 32]),
    ],
}
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({'output': str(OUTPUT_PATH), 'updatedAt': data['updatedAt'], 'groups': [(g['name'], len(g['players']), sum(len(v) for v in g['matches'].values())) for g in data['groups']]}, ensure_ascii=False))
